"""
ORB (Opening Range Breakout) 交易策略实现
根据配置文件读取 5 分钟 K 线数据
实现基于首个 5 分钟 K 线的交易策略
包含完整的风险控制和交易记录输出
自动检测并获取缺失数据
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from datetime import datetime, timedelta
import os
import matplotlib.dates as mdates
import matplotlib.ticker as mtick
import config  # 导入配置文件
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import sys
import importlib.util
from openpyxl.utils import get_column_letter

class ORB_Trading_Strategy:
    def __init__(self, symbol, start_date=None, end_date=None, initial_capital=None, auto_fetch_data=True):
        """
        初始化ORB策略
        
        参数:
        symbol (str): 交易品种代码
        start_date (str): 交易开始日期，格式'YYYY-MM-DD'，若为None则使用配置
        end_date (str): 交易结束日期，格式'YYYY-MM-DD'，若为None则使用配置
        initial_capital (float): 初始资金，若为None则使用配置
        auto_fetch_data (bool): 是否在数据缺失时自动获取
        """
        self.symbol = symbol
        self.start_date = start_date if start_date else config.START_DATE
        self.end_date = end_date if end_date else config.END_DATE
        self.initial_capital = initial_capital if initial_capital else config.INITIAL_CAPITAL
        self.current_capital = self.initial_capital
        self.auto_fetch_data = auto_fetch_data
        
        # 从配置文件获取策略参数
        self.commission_per_share = config.COMMISSION_PER_SHARE
        self.max_leverage = config.MAX_LEVERAGE  # 最大杠杆4倍
        self.risk_per_trade = config.RISK_PER_TRADE  # 每笔风险1%
        self.take_profit_r = config.TAKE_PROFIT_R  # 止盈目标为10R
        
        # 交易记录
        self.trades = []
        
        # 构建5分钟数据文件路径
        data_path = f"{config.DATA_DIR}/{self.symbol}_5min_full_{self.start_date}_to_{self.end_date}.csv"
        
        # 检查数据文件是否存在，如果不存在则尝试获取
        if not os.path.exists(data_path):
            if self.auto_fetch_data:
                print(f"5分钟数据文件不存在: {data_path}，尝试自动获取数据...")
                self._fetch_data_from_ibkr()
                
                # 再次检查数据文件是否存在
                if not os.path.exists(data_path):
                    raise FileNotFoundError(f"自动获取数据失败，数据文件仍不存在: {data_path}")
            else:
                raise FileNotFoundError(f"5分钟数据文件不存在: {data_path}，请先使用fetcher获取数据")
        
        # 读取5分钟数据
        self.data = self._load_market_data(data_path)
        
        # 处理结果
        self.trading_results = None
    
    def _fetch_data_from_ibkr(self):
        """从IBKR获取5分钟数据"""
        try:
            # 尝试导入数据获取模块
            spec = importlib.util.find_spec('ibkr_5min_data_fetcher')
            if spec is None:
                print("未找到ibkr_5min_data_fetcher模块，尝试从当前目录加载")
                
                # 如果模块不在路径中，尝试从当前目录加载
                if os.path.exists('ibkr_5min_data_fetcher.py'):
                    spec = importlib.util.spec_from_file_location("ibkr_5min_data_fetcher", "ibkr_5min_data_fetcher.py")
                    ibkr_fetcher = importlib.util.module_from_spec(spec)
                    spec.loader.exec_module(ibkr_fetcher)
                else:
                    raise ImportError("无法找到ibkr_5min_data_fetcher.py文件")
            else:
                ibkr_fetcher = importlib.import_module('ibkr_5min_data_fetcher')
            
            print(f"正在从IBKR获取 {self.symbol} 的5分钟数据...")
            
            # 调用数据获取函数
            result = ibkr_fetcher.get_symbol_5min_data(
                self.symbol, 
                self.start_date, 
                self.end_date,
                config.DATA_DIR
            )
            
            if result and result['full_data']['success']:
                print(f"成功获取 {self.symbol} 的5分钟数据：{result['full_data']['rows']} 行")
                return True
            else:
                error_msg = result['full_data']['error'] if result and 'full_data' in result else "未知错误"
                print(f"数据获取失败: {error_msg}")
                return False
                
        except Exception as e:
            print(f"获取数据时出错: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
    
    def _load_market_data(self, data_path):
        """加载5分钟市场数据"""
        print(f"加载 {self.symbol} 5分钟市场数据: {data_path}")
        
        data = pd.read_csv(data_path)
        
        # 查看数据中的date列格式
        try:
            print(f"Date column format sample: {data['date'].iloc[0]}")
            
            # 尝试强制将date列转换为datetime，忽略时区信息
            try:
                data['date'] = pd.to_datetime(data['date'], utc=True)
            except Exception as e:
                print(f"转换为时区感知的datetime失败: {str(e)}")
                try:
                    # 尝试不同的解析方法
                    data['date'] = pd.to_datetime(data['date'], errors='coerce')
                    # 检查是否有NaT值
                    if data['date'].isna().any():
                        print(f"警告: 转换后有 {data['date'].isna().sum()} 行日期值无效")
                except Exception as e:
                    print(f"转换日期列失败，请检查CSV文件中date列的格式: {str(e)}")
                    # 作为最后手段，将date作为字符串处理
                    data['date'] = data['date'].astype(str)
        
            # 添加日期列以便按天分组
            if pd.api.types.is_datetime64_any_dtype(data['date']):
                data['trade_date'] = data['date'].dt.date
            else:
                # 如果date不是datetime，尝试从字符串提取日期部分
                print("日期列不是datetime类型，尝试从字符串提取日期...")
                try:
                    # 假设date格式为 "YYYY-MM-DD HH:MM:SS"
                    data['trade_date'] = data['date'].str.split(' ').str[0]
                    data['trade_date'] = pd.to_datetime(data['trade_date']).dt.date
                except Exception as e:
                    print(f"提取日期失败: {str(e)}")
                    raise ValueError("无法处理日期列，请检查数据格式")
        
        except Exception as e:
            print(f"处理日期时出错: {str(e)}")
            import traceback
            traceback.print_exc()
            raise
        
        # 筛选日期范围
        start = pd.to_datetime(self.start_date).date()
        end = pd.to_datetime(self.end_date).date()
        data = data[(data['trade_date'] >= start) & (data['trade_date'] <= end)]
        
        if data.empty:
            raise ValueError(f"数据筛选后为空，请检查日期范围是否有效")
        
        # 确保数据按日期排序
        data = data.sort_values('date')
        
        # 验证交易日期是否按顺序排列
        sorted_dates = sorted(data['trade_date'].unique())
        actual_dates = list(data['trade_date'].unique())
        if sorted_dates != actual_dates:
            print("警告: 交易日期未按时间顺序排列! 正在重新排序...")
            # 已经在前面排序了，所以只需要警告
        
        # 可选：验证数据一致性
        self.validate_data_consistency(data)
        
        print(f"成功加载 {self.symbol} 5分钟市场数据，包含 {len(data['trade_date'].unique())} 个交易日")
        return data
    
    def validate_data_consistency(self, data):
        """验证数据的一致性"""
        print("验证数据一致性...")
        
        # 检查日期连续性
        trade_dates = sorted(data['trade_date'].unique())
        if len(trade_dates) < 2:
            print("警告: 交易日期少于2天")
            return
        
        # 检查相邻交易日的间隔
        date_diffs = [(trade_dates[i+1] - trade_dates[i]).days for i in range(len(trade_dates)-1)]
        max_gap = max(date_diffs)
        avg_gap = sum(date_diffs) / len(date_diffs)
        
        print(f"交易日期范围: {trade_dates[0]} 至 {trade_dates[-1]}")
        print(f"总交易日数: {len(trade_dates)}")
        print(f"平均间隔: {avg_gap:.2f} 天")
        print(f"最大间隔: {max_gap} 天")
        
        if max_gap > 5:  # 假设正常情况下周末最多3天间隔
            print(f"警告: 发现异常大的日期间隔 ({max_gap} 天)，可能存在数据缺失")
        
        # 检查每个交易日的K线数量
        bars_per_day = data.groupby('trade_date').size()
        min_bars = bars_per_day.min()
        max_bars = bars_per_day.max()
        avg_bars = bars_per_day.mean()
        
        print(f"每日平均K线数: {avg_bars:.2f}")
        print(f"最少K线数: {min_bars} (日期: {bars_per_day.idxmin()})")
        print(f"最多K线数: {max_bars} (日期: {bars_per_day.idxmax()})")
        
        if min_bars < 50:  # 假设正常交易日至少应有70多根5分钟K线
            print(f"警告: 某些交易日K线数量异常少 ({min_bars})，可能数据不完整")
        
        # 返回是否通过验证
        return max_gap <= 5 and min_bars >= 50
    
    def run_strategy(self):
        """运行ORB交易策略"""
        print(f"开始运行 {self.symbol} ORB交易策略 ({self.start_date} 至 {self.end_date})...")
        
        # 按交易日分组处理数据
        trade_dates = self.data['trade_date'].unique()
        
        # 调试信息
        print(f"总交易日数: {len(trade_dates)}")
        print(f"初始资金: {self.initial_capital}")
        print(f"当前资金: {self.current_capital}")
        
        for date in trade_dates:
            # 获取当天的数据
            day_data = self.data[self.data['trade_date'] == date].sort_values('date')
            
            # 确保当天至少有两根K线
            if len(day_data) < 2:
                print(f"警告: {date} 数据不足，跳过")
                continue
            
            # 提取前两根5分钟K线
            first_candle = day_data.iloc[0]
            second_candle = day_data.iloc[1]
            
            # 跳过十字星日
            if abs(first_candle['close'] - first_candle['open']) < 0.0001:
                print(f"{date}: 第一根K线为十字星，跳过交易")
                continue
            
            # 确定交易方向
            direction = 1 if first_candle['close'] > first_candle['open'] else -1
            
            # 入场价格 (第二根5分钟K线开盘价)
            entry_price = second_candle['open']
            
            # 止损价格 (第一根5分钟K线的高/低点)
            stop_loss_price = first_candle['low'] if direction == 1 else first_candle['high']
            
            # 计算风险
            risk = abs(entry_price - stop_loss_price)
            if risk < 0.0001:  # 防止除以零
                print(f"{date}: 风险太小，跳过交易")
                continue
            
            # 设置止盈目标 (10R)
            take_profit_price = entry_price + (self.take_profit_r * risk * direction)
            
            # 计算持仓规模
            shares = self._calculate_position_size(entry_price, stop_loss_price)
            
            # 如果持仓规模为0，跳过此次交易
            if shares == 0:
                print(f"{date}: 计算持仓为0，跳过交易")
                continue
            
            # 计算佣金
            commission = shares * self.commission_per_share
            
            # 提取当天剩余K线数据，用于模拟交易过程
            remaining_candles = day_data.iloc[2:]
            
            # 初始化出场变量
            exit_price = None
            exit_reason = None
            exit_time = None
            
            # 遍历剩余K线，检查是否触及止损或止盈
            for _, candle in remaining_candles.iterrows():
                # 做多情况
                if direction == 1:
                    # 检查是否触及止损
                    if candle['low'] <= stop_loss_price:
                        exit_price = stop_loss_price
                        exit_reason = 'Stop Loss'
                        exit_time = candle['date']
                        break
                    # 检查是否触及止盈
                    elif candle['high'] >= take_profit_price:
                        exit_price = take_profit_price
                        exit_reason = 'Take Profit'
                        exit_time = candle['date']
                        break
                # 做空情况
                else:
                    # 检查是否触及止损
                    if candle['high'] >= stop_loss_price:
                        exit_price = stop_loss_price
                        exit_reason = 'Stop Loss'
                        exit_time = candle['date']
                        break
                    # 检查是否触及止盈
                    elif candle['low'] <= take_profit_price:
                        exit_price = take_profit_price
                        exit_reason = 'Take Profit'
                        exit_time = candle['date']
                        break
            
            # 如果没有触及止损或止盈，则以收盘价平仓
            if exit_price is None:
                last_candle = day_data.iloc[-1]
                exit_price = last_candle['close']
                exit_reason = 'End of Day'
                exit_time = last_candle['date']
            
            # 计算交易盈亏
            trade_pnl = (exit_price - entry_price) * direction * shares - commission
            pnl_pct = trade_pnl / self.current_capital
            pnl_in_r = ((exit_price - entry_price) * direction) / risk
            
            # 记录交易
            trade = {
                'date': day_data.iloc[0]['date'],
                'entry_time': second_candle['date'],
                'exit_time': exit_time,
                'direction': direction,
                'entry_price': entry_price,
                'stop_loss': stop_loss_price,
                'take_profit': take_profit_price,
                'exit_price': exit_price,
                'exit_reason': exit_reason,
                'shares': shares,
                'risk_amount': self.current_capital * self.risk_per_trade,
                'risk_in_points': risk,
                'commission': commission,
                'pnl': trade_pnl,
                'pnl_pct': pnl_pct,
                'pnl_in_r': pnl_in_r
            }
            
            # 更新账户资金
            self.current_capital += trade_pnl
            
            # 添加交易记录
            self.trades.append(trade)
            
            # 打印交易摘要 (增加更多信息)
            trade_direction = "多" if direction == 1 else "空"
            print(f"{date}: {trade_direction}头交易, 出场原因: {exit_reason}, 盈亏: {pnl_in_r:.2f}R (${trade_pnl:.2f}), 当前资金: ${self.current_capital:.2f}")
        
        # 交易结束，打印总体摘要
        print(f"\n交易结束，总交易次数: {len(self.trades)}")
        print(f"初始资金: ${self.initial_capital:.2f}")
        print(f"最终资金: ${self.current_capital:.2f}")
        print(f"总收益: ${self.current_capital - self.initial_capital:.2f} ({(self.current_capital/self.initial_capital - 1)*100:.2f}%)")
        
        # 计算交易结果
        self._calculate_trading_results()
        
        return self.trading_results
    
    def _calculate_position_size(self, entry_price, stop_loss_price):
        """
        计算持仓规模，根据公式:
        Shares = int[min((A * 0.01 / R), (4 * A / P))]
        
        A: 账户资金
        R: 入场价与止损价的差距
        P: 入场价格
        """
        risk = abs(entry_price - stop_loss_price)
        risk_amount = self.current_capital * self.risk_per_trade
        
        # 基于风险的头寸大小
        risk_based_shares = int(risk_amount / risk)
        
        # 基于杠杆的头寸大小
        leverage_based_shares = int((self.max_leverage * self.current_capital) / entry_price)
        
        # 取两者的较小值
        shares = min(risk_based_shares, leverage_based_shares)
        
        return shares
    
    def _calculate_trading_results(self):
        """计算交易结果统计数据"""
        if not self.trades:
            self.trading_results = {
                'symbol': self.symbol,
                'total_trades': 0,
                'profitable_trades': 0,
                'win_rate': 0,
                'total_pnl': 0,
                'total_pnl_pct': 0,
                'avg_pnl_in_r': 0,
                'initial_capital': self.initial_capital,
                'final_capital': self.current_capital,
                'trades': []
            }
            return
        
        # 转换交易记录为DataFrame以便分析，并确保按日期排序
        trades_df = pd.DataFrame(self.trades)
        # 确保日期列为datetime类型且排序正确
        if not pd.api.types.is_datetime64_any_dtype(trades_df['date']):
            trades_df['date'] = pd.to_datetime(trades_df['date'])
        trades_df = trades_df.sort_values('date')
        
        # 验证交易记录
        print("\n验证交易记录:")
        print(f"交易记录总数: {len(trades_df)}")
        print(f"总盈亏金额: ${trades_df['pnl'].sum():.2f}")
        print(f"最大单笔盈利: ${trades_df['pnl'].max():.2f}")
        print(f"最大单笔亏损: ${trades_df['pnl'].min():.2f}")
        
        # 计算基本统计数据
        total_trades = len(trades_df)
        profitable_trades = len(trades_df[trades_df['pnl'] > 0])
        win_rate = profitable_trades / total_trades
        
        total_pnl = trades_df['pnl'].sum()
        total_pnl_pct = (self.current_capital - self.initial_capital) / self.initial_capital
        
        # 按出场原因分组统计
        exit_reason_stats = trades_df.groupby('exit_reason').agg({
            'pnl': ['count', 'sum', 'mean'],
            'pnl_in_r': 'mean'
        })
        
        # 存储交易结果
        self.trading_results = {
            'symbol': self.symbol,
            'total_trades': total_trades,
            'profitable_trades': profitable_trades,
            'win_rate': win_rate,
            'total_pnl': total_pnl,
            'total_pnl_pct': total_pnl_pct,
            'avg_pnl_in_r': trades_df['pnl_in_r'].mean(),
            'exit_reason_stats': exit_reason_stats,
            'initial_capital': self.initial_capital,
            'final_capital': self.current_capital,
            'trades': trades_df
        }
        
        # 打印结果摘要
        print("\n交易结果摘要:")
        print(f"总交易次数: {total_trades}")
        print(f"盈利交易: {profitable_trades} ({win_rate:.2%})")
        print(f"总盈亏: ${total_pnl:.2f} ({total_pnl_pct:.2%})")
        print(f"平均盈亏(R): {self.trading_results['avg_pnl_in_r']:.2f}R")
        print(f"最终资金: ${self.current_capital:.2f}")
    
    def generate_excel_report(self, save_path=None):
        """将交易结果导出为Excel表格"""
        if not self.trading_results:
            print("没有交易结果，请先运行策略")
            return
        
        if save_path is None:
            output_dir = f"{config.OUTPUT_DIR}/{self.symbol}"
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)
            save_path = f"{output_dir}/{self.symbol}_ORB_Trading_Results.xlsx"
        
        print(f"正在生成Excel报告: {save_path}...")
        
        # 创建Excel工作簿和工作表
        wb = openpyxl.Workbook()
        
        # 创建摘要工作表
        summary_sheet = wb.active
        summary_sheet.title = "Trading Summary"
        
        # 添加标题
        summary_sheet['A1'] = f"{self.symbol} ORB Trading Strategy Results"
        summary_sheet['A1'].font = Font(bold=True, size=14)
        summary_sheet.merge_cells('A1:F1')
        
        # 添加基本信息
        summary_sheet['A3'] = "Trading Period:"
        summary_sheet['B3'] = f"{self.start_date} to {self.end_date}"
        
        summary_sheet['A4'] = "Initial Capital:"
        summary_sheet['B4'] = f"${self.initial_capital:,.2f}"
        
        summary_sheet['A5'] = "Final Capital:"
        summary_sheet['B5'] = f"${self.trading_results['final_capital']:,.2f}"
        
        summary_sheet['A6'] = "Total P&L:"
        summary_sheet['B6'] = f"${self.trading_results['total_pnl']:,.2f}"
        summary_sheet['C6'] = f"({self.trading_results['total_pnl_pct']:.2%})"
        
        # 添加交易统计
        summary_sheet['A8'] = "Trading Statistics"
        summary_sheet['A8'].font = Font(bold=True, size=12)
        summary_sheet.merge_cells('A8:F8')
        
        summary_sheet['A9'] = "Total Trades:"
        summary_sheet['B9'] = self.trading_results['total_trades']
        
        summary_sheet['A10'] = "Profitable Trades:"
        summary_sheet['B10'] = self.trading_results['profitable_trades']
        
        summary_sheet['A11'] = "Win Rate:"
        summary_sheet['B11'] = f"{self.trading_results['win_rate']:.2%}"
        
        summary_sheet['A12'] = "Average P&L in R:"
        summary_sheet['B12'] = f"{self.trading_results['avg_pnl_in_r']:.2f}R"
        
        # 按出场原因分析
        summary_sheet['A14'] = "Exit Reason Analysis"
        summary_sheet['A14'].font = Font(bold=True, size=12)
        summary_sheet.merge_cells('A14:F14')
        
        summary_sheet['A15'] = "Exit Reason"
        summary_sheet['B15'] = "Count"
        summary_sheet['C15'] = "Total P&L"
        summary_sheet['D15'] = "Avg P&L"
        summary_sheet['E15'] = "Avg R Multiple"
        
        # 设置标题行样式
        for col in range(1, 6):
            cell = summary_sheet.cell(row=15, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # 添加出场原因统计数据
        row = 16
        for exit_reason, stats in self.trading_results['exit_reason_stats'].iterrows():
            summary_sheet[f'A{row}'] = exit_reason
            summary_sheet[f'B{row}'] = int(stats[('pnl', 'count')])
            summary_sheet[f'C{row}'] = f"${stats[('pnl', 'sum')]:,.2f}"
            summary_sheet[f'D{row}'] = f"${stats[('pnl', 'mean')]:,.2f}"
            summary_sheet[f'E{row}'] = f"{stats[('pnl_in_r', 'mean')]:,.2f}R"
            row += 1
        
        # 创建详细交易记录工作表
        trades_sheet = wb.create_sheet("Trade Details")
        
        # 添加标题行
        headers = [
            "Date", "Entry Time", "Exit Time", "Direction", 
            "Entry Price", "Stop Loss", "Take Profit", "Exit Price", 
            "Exit Reason", "Shares", "Risk ($)", "Risk (pts)", 
            "Commission", "P&L ($)", "P&L (%)", "P&L in R"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = trades_sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # 添加交易数据
        trades_data = self.trading_results['trades']
        
        for row_idx, trade in enumerate(trades_data.itertuples(), 2):
            # 日期和时间
            try:
                # 尝试将日期对象格式化为字符串
                if isinstance(trade.date, datetime) or hasattr(trade.date, 'strftime'):
                    date_str = trade.date.strftime('%Y-%m-%d')
                elif isinstance(trade.date, str):
                    # 如果已经是字符串，尝试解析并重新格式化
                    try:
                        date_str = pd.to_datetime(trade.date).strftime('%Y-%m-%d')
                    except:
                        # 如果无法解析，则直接使用
                        date_str = trade.date.split(' ')[0] if ' ' in trade.date else trade.date
                else:
                    # 其他情况，转为字符串
                    date_str = str(trade.date)
                
                trades_sheet.cell(row=row_idx, column=1).value = date_str
                
                # 处理入场时间和出场时间
                for col_idx, time_value in [(2, trade.entry_time), (3, trade.exit_time)]:
                    try:
                        if isinstance(time_value, datetime) or hasattr(time_value, 'strftime'):
                            time_str = time_value.strftime('%H:%M:%S')
                        elif isinstance(time_value, str):
                            # 尝试提取时间部分
                            if ' ' in time_value:
                                time_str = time_value.split(' ')[1]
                            else:
                                time_str = time_value
                        else:
                            time_str = str(time_value)
                        
                        trades_sheet.cell(row=row_idx, column=col_idx).value = time_str
                    except Exception as e:
                        trades_sheet.cell(row=row_idx, column=col_idx).value = str(time_value)
                        print(f"格式化时间 {time_value} 时出错: {str(e)}")
            except Exception as e:
                print(f"处理交易日期时间时出错: {str(e)}")
                trades_sheet.cell(row=row_idx, column=1).value = str(trade.date)
                trades_sheet.cell(row=row_idx, column=2).value = str(trade.entry_time)
                trades_sheet.cell(row=row_idx, column=3).value = str(trade.exit_time)
            
            # 方向
            direction_text = "Long" if trade.direction == 1 else "Short"
            trades_sheet.cell(row=row_idx, column=4).value = direction_text
            
            # 价格
            trades_sheet.cell(row=row_idx, column=5).value = trade.entry_price
            trades_sheet.cell(row=row_idx, column=6).value = trade.stop_loss
            trades_sheet.cell(row=row_idx, column=7).value = trade.take_profit
            trades_sheet.cell(row=row_idx, column=8).value = trade.exit_price
            
            # 其他信息
            trades_sheet.cell(row=row_idx, column=9).value = trade.exit_reason
            trades_sheet.cell(row=row_idx, column=10).value = trade.shares
            trades_sheet.cell(row=row_idx, column=11).value = trade.risk_amount
            trades_sheet.cell(row=row_idx, column=12).value = trade.risk_in_points
            trades_sheet.cell(row=row_idx, column=13).value = trade.commission
            
            # 盈亏
            pnl_cell = trades_sheet.cell(row=row_idx, column=14)
            pnl_cell.value = trade.pnl
            
            # 设置盈亏单元格颜色
            if trade.pnl > 0:
                pnl_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif trade.pnl < 0:
                pnl_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            trades_sheet.cell(row=row_idx, column=15).value = trade.pnl_pct
            
            # R倍数
            r_cell = trades_sheet.cell(row=row_idx, column=16)
            r_cell.value = trade.pnl_in_r
            
            # 设置R倍数单元格颜色
            if trade.pnl_in_r > 0:
                r_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif trade.pnl_in_r < 0:
                r_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # 自动调整列宽
        for sheet in [summary_sheet, trades_sheet]:
            for col in sheet.columns:
                max_length = 0
                # 获取列字母，避免使用合并单元格
                column = get_column_letter(col[0].column)
                for cell in col:
                    # 跳过合并单元格
                    if isinstance(cell, openpyxl.cell.cell.MergedCell):
                        continue
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column].width = adjusted_width
        
        # 创建资金曲线工作表
        equity_sheet = wb.create_sheet("Equity Curve")
        
        # 计算每日权益
        trades_df = self.trading_results['trades'].copy()
        if not trades_df.empty:
            # 确保日期列是datetime类型
            try:
                if not pd.api.types.is_datetime64_any_dtype(trades_df['date']):
                    print("将日期列转换为datetime格式以进行权益曲线计算...")
                    # 尝试不同方法转换日期
                    try:
                        trades_df['date'] = pd.to_datetime(trades_df['date'], errors='coerce')
                    except Exception as e:
                        print(f"标准日期转换失败: {str(e)}")
                        # 尝试从字符串中提取日期
                        try:
                            if trades_df['date'].dtype == object:  # 如果是字符串
                                # 尝试提取 YYYY-MM-DD 格式
                                trades_df['date'] = trades_df['date'].astype(str).str.split(' ').str[0]
                                trades_df['date'] = pd.to_datetime(trades_df['date'], errors='coerce')
                        except Exception as e2:
                            print(f"日期转换失败，权益曲线可能不准确: {str(e2)}")
                
                # 检查是否有缺失日期
                if trades_df['date'].isna().any():
                    print(f"警告: 有 {trades_df['date'].isna().sum()} 行的日期值无效，将被排除")
                    trades_df = trades_df.dropna(subset=['date'])
                
                # 确保日期按升序排列 - 这是解决问题的关键！
                trades_df = trades_df.sort_values('date')
                
                # 获取第一天和最后一天的日期
                first_date = trades_df['date'].min()
                last_date = trades_df['date'].max()
                
                # 创建完整的日期范围
                date_range = pd.date_range(start=first_date, end=last_date)
                equity_df = pd.DataFrame(index=date_range)
                equity_df.index.name = 'date'
                equity_df['equity'] = np.nan
                
                # 设置初始资金
                equity_df.loc[equity_df.index[0], 'equity'] = self.initial_capital
                
                # 计算每日权益 - 完全重写这部分逻辑
                print("计算权益曲线...")
                
                # 1. 获取每日交易盈亏
                daily_pnl = {}
                for date, group in trades_df.groupby(trades_df['date'].dt.date):
                    daily_pnl[date] = group['pnl'].sum()
                
                # 2. 按时间顺序计算累积盈亏
                cumulative_pnl = 0
                
                # 3. 为每一个日期索引计算权益
                for i, date in enumerate(equity_df.index):
                    date_only = date.date()  # 提取日期部分，不含时间
                    
                    # 如果是第一天，设置为初始资金
                    if i == 0:
                        equity_df.loc[date, 'equity'] = self.initial_capital
                        continue
                    
                    # 检查当天是否有交易
                    if date_only in daily_pnl:
                        # 当天有交易，累加盈亏
                        cumulative_pnl += daily_pnl[date_only]
                        print(f"日期 {date_only}: 当日盈亏 = {daily_pnl[date_only]:.2f}, 累计盈亏 = {cumulative_pnl:.2f}")
                    
                    # 更新权益值 = 初始资金 + 累计盈亏
                    equity_df.loc[date, 'equity'] = self.initial_capital + cumulative_pnl
                
                # 打印验证信息
                print(f"权益曲线第一天: {equity_df['equity'].iloc[0]:.2f}")
                print(f"权益曲线最后一天: {equity_df['equity'].iloc[-1]:.2f}")
                print(f"总盈亏: {cumulative_pnl:.2f}")
                
                # 添加 Buy & Hold 策略
                # 获取每天的市场数据
                market_data = self.data.copy()
                try:
                    # 确保市场数据的日期列格式正确
                    if isinstance(market_data['trade_date'].iloc[0], (str, object)):
                        market_data['date'] = pd.to_datetime(market_data['trade_date'])
                    else:
                        market_data['date'] = market_data['trade_date']
                except Exception as e:
                    print(f"处理市场数据日期时出错: {str(e)}")
                    market_data['date'] = pd.to_datetime(market_data['trade_date'], errors='coerce')
                
                # 确保市场数据按日期排序
                market_data = market_data.sort_values('date')
                
                # 获取每天的收盘价
                daily_close = market_data.groupby('trade_date')['close'].last().reset_index()
                daily_close['date'] = pd.to_datetime(daily_close['trade_date'])
                daily_close = daily_close.sort_values('date')  # 确保按日期排序
                daily_close = daily_close.set_index('date')
                
                # 在处理时区不一致的情况
                try:
                    # 安全方式合并数据
                    # 转换为字符串日期进行合并，避免时区问题
                    temp_equity_df = equity_df.reset_index()
                    temp_equity_df['date_str'] = temp_equity_df['date'].dt.strftime('%Y-%m-%d')
                    temp_equity_df = temp_equity_df.set_index('date_str')
                    
                    temp_daily_close = daily_close.reset_index()
                    temp_daily_close['date_str'] = temp_daily_close['date'].dt.strftime('%Y-%m-%d')
                    temp_daily_close = temp_daily_close.set_index('date_str')
                    
                    # 使用字符串索引合并
                    combined_df = temp_equity_df.join(temp_daily_close[['close']], how='left')
                    combined_df = combined_df.sort_index()  # 确保仍按日期排序
                    
                    # 恢复原始索引
                    combined_df = combined_df.reset_index().set_index('date')
                    equity_df = combined_df[['equity', 'close']]
                except Exception as e:
                    print(f"合并数据时出错: {str(e)}")
                    # 使用备选方法
                    
                # 前向填充收盘价的缺失值
                equity_df['close'] = equity_df['close'].ffill()
                
                # 计算初始可买入的股票数量
                initial_close = equity_df['close'].iloc[0]
                if not pd.isna(initial_close) and initial_close > 0:
                    shares_held = self.initial_capital / initial_close
                    equity_df['buy_hold'] = equity_df['close'] * shares_held
                else:
                    print("警告: 无法计算Buy & Hold策略，初始收盘价无效")
                    equity_df['buy_hold'] = np.nan
                
                # 最后确保按日期排序
                equity_df = equity_df.sort_index()
                
                # 添加数据到工作表
                equity_sheet['A1'] = "Date"
                equity_sheet['B1'] = "ORB Strategy"
                equity_sheet['C1'] = "Buy & Hold"
                equity_sheet['D1'] = "Relative Performance"
                
                # 设置标题行样式
                for col in range(1, 5):
                    cell = equity_sheet.cell(row=1, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                
                # 使用已排序的equity_df填充表格
                for i, (date, row) in enumerate(equity_df.iterrows(), 2):
                    equity_sheet[f'A{i}'] = date.strftime('%Y-%m-%d')
                    
                    # 直接写入ORB策略的权益值
                    if pd.notna(row['equity']):
                        equity_sheet[f'B{i}'] = row['equity']
                    else:
                        # 如果没有数据，使用前一天的值或初始资金
                        equity_sheet[f'B{i}'] = self.initial_capital if i == 2 else equity_sheet[f'B{i-1}'].value
                    
                    # Buy & Hold 值
                    if 'buy_hold' in row and pd.notna(row['buy_hold']):
                        equity_sheet[f'C{i}'] = row['buy_hold']
                    else:
                        # 如果没有数据，使用前一天的值或初始资金
                        equity_sheet[f'C{i}'] = self.initial_capital if i == 2 else equity_sheet[f'C{i-1}'].value
                    
                    # 计算相对表现 (ORB vs Buy & Hold)
                    try:
                        b_value = equity_sheet[f'B{i}'].value
                        c_value = equity_sheet[f'C{i}'].value
                        
                        if b_value is not None and c_value is not None and c_value != 0:
                            relative_perf = (b_value / c_value - 1) * 100
                            equity_sheet[f'D{i}'] = f"{relative_perf:.2f}%"
                            
                            # 设置相对表现单元格颜色
                            perf_cell = equity_sheet.cell(row=i, column=4)
                            if relative_perf > 0:
                                perf_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            elif relative_perf < 0:
                                perf_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    except Exception as e:
                        print(f"计算相对表现时出错: {str(e)}")
                
                # 创建权益曲线图表
                try:
                    from openpyxl.chart import LineChart, Reference
                    
                    chart = LineChart()
                    chart.title = f"{self.symbol} ORB Strategy vs Buy & Hold"
                    chart.style = 2
                    chart.x_axis.title = "Date"
                    chart.y_axis.title = "Portfolio Value ($)"
                    
                    # 添加数据
                    data = Reference(equity_sheet, min_col=2, min_row=1, max_col=3, max_row=len(equity_df)+1)
                    chart.add_data(data, titles_from_data=True)
                    
                    # 添加日期轴
                    dates = Reference(equity_sheet, min_col=1, min_row=2, max_row=len(equity_df)+1)
                    chart.set_categories(dates)
                    
                    # 将图表添加到工作表
                    equity_sheet.add_chart(chart, "F5")
                except Exception as e:
                    print(f"创建图表时出错: {str(e)}")
            except Exception as e:
                print(f"生成权益曲线时出错: {str(e)}")
                import traceback
                traceback.print_exc()
                equity_sheet['A1'] = "生成权益曲线时发生错误"
                equity_sheet['A2'] = str(e)
        else:
            # 如果没有交易数据，添加一个简单的消息
            equity_sheet['A1'] = "No trading data available for equity curve"
        
        # 保存Excel文件
        wb.save(save_path)
        print(f"Excel报告已生成: {save_path}")
        
        return save_path
    
    def generate_summary_report(self, save_path=None):
        """生成摘要报告"""
        if not self.trading_results:
            print("没有交易结果，请先运行策略")
            return
        
        if save_path is None:
            output_dir = f"{config.OUTPUT_DIR}/{self.symbol}"
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)
            save_path = f"{output_dir}/{self.symbol}_ORB_Summary.txt"
        
        report = f"""
=============================================
     {self.symbol} ORB Trading Strategy Summary
=============================================

Trading Period: {self.start_date} to {self.end_date}
Initial Capital: ${self.initial_capital:.2f}
Final Capital: ${self.trading_results['final_capital']:.2f}

---------------------------------------------
            Trading Statistics
---------------------------------------------
Total Trades: {self.trading_results['total_trades']}
Profitable Trades: {self.trading_results['profitable_trades']}
Win Rate: {self.trading_results['win_rate']:.2%}

Total P&L: ${self.trading_results['total_pnl']:.2f} ({self.trading_results['total_pnl_pct']:.2%})
Average P&L in Risk Units (R): {self.trading_results['avg_pnl_in_r']:.2f}R

---------------------------------------------
          Exit Reason Analysis
---------------------------------------------
"""
        
        # 添加出场原因统计
        exit_stats = self.trading_results['exit_reason_stats']
        for exit_reason, stats in exit_stats.iterrows():
            report += f"{exit_reason}: {int(stats[('pnl', 'count')])} trades, "
            report += f"Total P&L: ${stats[('pnl', 'sum')]:.2f}, "
            report += f"Avg P&L: ${stats[('pnl', 'mean')]:.2f}, "
            report += f"Avg R: {stats[('pnl_in_r', 'mean')]:.2f}R\n"
        
        report += "\n=============================================\n"
        
        # 保存报告
        with open(save_path, 'w') as f:
            f.write(report)
        
        print(f"摘要报告已保存至: {save_path}")
        return report


def run_trading_strategy(symbol, start_date=None, end_date=None, initial_capital=None):
    """运行单个交易品种的ORB策略"""
    # 使用配置文件中的参数，除非明确指定
    start_date = start_date or config.START_DATE
    end_date = end_date or config.END_DATE
    initial_capital = initial_capital or config.INITIAL_CAPITAL
    
    # 创建输出目录
    output_dir = f"{config.OUTPUT_DIR}/{symbol}"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    try:
        # 创建并运行ORB交易策略
        strategy = ORB_Trading_Strategy(symbol, start_date, end_date, initial_capital)
        results = strategy.run_strategy()
        
        # 生成Excel报告
        excel_path = f"{output_dir}/{symbol}_ORB_Trading_Results.xlsx"
        strategy.generate_excel_report(save_path=excel_path)
        
        # 生成摘要报告
        summary_path = f"{output_dir}/{symbol}_ORB_Summary.txt"
        summary = strategy.generate_summary_report(save_path=summary_path)
        print(summary)
        
        return {
            'strategy': strategy,
            'results': results,
            'excel_path': excel_path,
            'summary_path': summary_path
        }
    
    except Exception as e:
        print(f"运行 {symbol} 策略时出错: {str(e)}")
        import traceback
        traceback.print_exc()
        return None


def run_all_symbols():
    """运行所有配置的交易品种策略"""
    results = {}
    
    for symbol in config.SYMBOLS:
        print(f"\n{'='*50}")
        print(f"开始运行 {symbol} ORB交易策略")
        print(f"{'='*50}")
        
        result = run_trading_strategy(symbol)
        results[symbol] = result
    
    return results


def main():
    """主函数"""
    print(f"ORB交易策略系统")
    print(f"交易周期: {config.START_DATE} 至 {config.END_DATE}")
    print(f"交易品种: {', '.join(config.SYMBOLS)}")
    print(f"初始资金: ${config.INITIAL_CAPITAL:.2f}")
    
    # 创建输出主目录
    if not os.path.exists(config.OUTPUT_DIR):
        os.makedirs(config.OUTPUT_DIR)
    
    # 运行所有品种策略
    results = run_all_symbols()
    
    print(f"\n所有策略运行完成！详细结果已保存到 {config.OUTPUT_DIR} 目录")


if __name__ == "__main__":
    main() 