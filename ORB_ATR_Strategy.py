"""
ORB_ATR_Strategy.py

基于 ORB 策略的改进版，使用 ATR 动态止损 + EoD 止盈
相比原始 ORB 策略，这个版本使用 ATR 计算止损距离，并取消止盈位，改为收盘平仓
这个改进在论文中显示可以显著提高策略的表现
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import os
import datetime
from datetime import timedelta
import matplotlib.dates as mdates
import matplotlib.ticker as mtick
import config
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import sys
import importlib.util
from openpyxl.utils import get_column_letter
from market_calendar import MarketCalendar

class ORB_ATR_Strategy:
    def __init__(self, symbol, start_date=None, end_date=None, initial_capital=None, auto_fetch_data=True, atr_period=14, atr_multiplier=0.05):
        """
        初始化ORB ATR策略
        
        参数:
        symbol (str): 交易品种代码
        start_date (str): 交易开始日期，格式'YYYY-MM-DD'，若为None则使用配置
        end_date (str): 交易结束日期，格式'YYYY-MM-DD'，若为None则使用配置
        initial_capital (float): 初始资金，若为None则使用配置
        auto_fetch_data (bool): 是否在数据缺失时自动获取
        atr_period (int): ATR计算周期，默认14天
        atr_multiplier (float): ATR乘数用于计算止损距离，默认0.05（5%）
        """
        self.symbol = symbol
        self.start_date = start_date if start_date else config.START_DATE
        self.end_date = end_date if end_date else config.END_DATE
        self.initial_capital = initial_capital if initial_capital else config.INITIAL_CAPITAL
        self.current_capital = self.initial_capital
        self.auto_fetch_data = auto_fetch_data
        
        # ATR 参数
        self.atr_period = atr_period
        self.atr_multiplier = atr_multiplier
        
        # 从配置文件获取策略参数
        self.commission_per_share = config.COMMISSION_PER_SHARE
        self.max_leverage = config.MAX_LEVERAGE  # 最大杠杆4倍
        self.risk_per_trade = config.RISK_PER_TRADE  # 每笔风险1%
        
        # 交易记录
        self.trades = []
        
        # 初始化市场日历
        self.market_calendar = MarketCalendar(config.TIMEZONE)
        
        # 构建5分钟数据文件路径
        data_path = f"{config.DATA_DIR}/{self.symbol}_5min_full_{self.start_date}_to_{self.end_date}.csv"
        
        # 检查数据文件是否存在，如果不存在则尝试获取
        if not os.path.exists(data_path):
            if self.auto_fetch_data:
                print(f"5分钟数据文件不存在: {data_path}，尝试自动获取数据...")
                self._fetch_data_from_ibkr()
                
                # 再次检查数据文件是否存在
                if not os.path.exists(data_path):
                    raise FileNotFoundError(f"自动获取数据失败，数据文件仍不存在: {data_path}")
            else:
                raise FileNotFoundError(f"5分钟数据文件不存在: {data_path}，请先使用fetcher获取数据")
        
        # 读取5分钟数据
        self.data = self._load_market_data(data_path)
        
        # 计算日线数据和ATR
        self.daily_data = self._calculate_daily_data()
        
        # 处理结果
        self.trading_results = None
    
    def _fetch_data_from_ibkr(self):
        """从IBKR获取5分钟数据"""
        try:
            # 直接导入和使用现有模块
            import ibkr_5min_data_fetcher
            
            print(f"正在从IBKR获取 {self.symbol} 的5分钟数据...")
            
            # 直接调用现有函数
            result = ibkr_5min_data_fetcher.get_symbol_5min_data(
                self.symbol, 
                self.start_date, 
                self.end_date,
                config.DATA_DIR
            )
            
            if result and result['full_data']['success']:
                print(f"成功获取 {self.symbol} 的5分钟数据：{result['full_data']['rows']} 行")
                return True
            else:
                error_msg = result['full_data']['error'] if result and 'full_data' in result else "未知错误"
                print(f"数据获取失败: {error_msg}")
                return False
                
        except Exception as e:
            print(f"获取数据时出错: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
    
    def _load_market_data(self, data_path):
        """加载5分钟市场数据"""
        print(f"加载 {self.symbol} 5分钟市场数据: {data_path}")
        
        data = pd.read_csv(data_path)
        
        # 查看数据中的date列格式
        try:
            print(f"Date column format sample: {data['date'].iloc[0]}")
            
            # 尝试强制将date列转换为datetime，处理混合时区
            try:
                # 使用utc=True参数处理混合时区
                data['date'] = pd.to_datetime(data['date'], utc=True)
            except Exception as e:
                print(f"转换为时区感知的datetime失败: {str(e)}")
                try:
                    # 尝试不同的解析方法
                    data['date'] = pd.to_datetime(data['date'], errors='coerce')
                    # 检查是否有NaT值
                    if data['date'].isna().any():
                        print(f"警告: 转换后有 {data['date'].isna().sum()} 行日期值无效")
                except Exception as e:
                    print(f"转换日期列失败，请检查CSV文件中date列的格式: {str(e)}")
                    # 作为最后手段，将date作为字符串处理
                    data['date'] = data['date'].astype(str)
        
            # 添加日期列以便按天分组
            if pd.api.types.is_datetime64_any_dtype(data['date']):
                data['trade_date'] = data['date'].dt.date
            else:
                # 如果date不是datetime，尝试从字符串提取日期部分
                print("日期列不是datetime类型，尝试从字符串提取日期...")
                try:
                    # 假设date格式为 "YYYY-MM-DD HH:MM:SS"
                    data['trade_date'] = data['date'].str.split(' ').str[0]
                    data['trade_date'] = pd.to_datetime(data['trade_date']).dt.date
                except Exception as e:
                    print(f"提取日期失败: {str(e)}")
                    raise ValueError("无法处理日期列，请检查数据格式")
        
        except Exception as e:
            print(f"处理日期时出错: {str(e)}")
            import traceback
            traceback.print_exc()
            raise
        
        # 筛选日期范围
        start_date = pd.to_datetime(self.start_date)
        end_date = pd.to_datetime(self.end_date)
        
        # 获取交易日列表，使用 datetime 对象而非 date 对象
        try:
            trading_days = self.market_calendar.get_trading_days(start_date, end_date)
            # 将结果转换为 date 对象列表以便后续匹配
            trading_days = [d.date() if hasattr(d, 'date') else d for d in trading_days]
            data = data[data['trade_date'].isin(trading_days)]
        except Exception as e:
            print(f"获取交易日列表时出错: {e}")
            # 备选方案：如果无法获取交易日列表，直接使用日期范围筛选
            start = pd.to_datetime(self.start_date).date()
            end = pd.to_datetime(self.end_date).date()
            data = data[(data['trade_date'] >= start) & (data['trade_date'] <= end)]
        
        if data.empty:
            raise ValueError(f"数据筛选后为空，请检查日期范围是否有效")
        
        print(f"成功加载 {self.symbol} 5分钟市场数据，包含 {len(data['trade_date'].unique())} 个交易日")
        return data
    
    def _calculate_daily_data(self):
        """计算日线数据和ATR指标"""
        print(f"计算 {self.symbol} 日线数据和ATR指标...")
        
        # 创建日线数据
        daily_data = self.data.groupby('trade_date').agg({
            'open': 'first',
            'high': 'max',
            'low': 'min',
            'close': 'last'
        }).reset_index()
        
        # 确保日期类型正确
        daily_data['trade_date'] = pd.to_datetime(daily_data['trade_date'])
        daily_data = daily_data.sort_values('trade_date')
        
        # 计算真实波幅 (True Range)
        daily_data['prev_close'] = daily_data['close'].shift(1)
        daily_data['tr1'] = daily_data['high'] - daily_data['low']
        daily_data['tr2'] = abs(daily_data['high'] - daily_data['prev_close'])
        daily_data['tr3'] = abs(daily_data['low'] - daily_data['prev_close'])
        daily_data['tr'] = daily_data[['tr1', 'tr2', 'tr3']].max(axis=1)
        
        # 计算ATR (N日平均真实波幅)
        daily_data['atr'] = daily_data['tr'].rolling(window=self.atr_period).mean()
        
        # 处理NaN值
        daily_data = daily_data.dropna(subset=['atr']).copy()
        
        print(f"成功计算ATR指标，剩余 {len(daily_data)} 个交易日")
        return daily_data
    
    def _get_atr_for_date(self, date):
        """获取指定日期的ATR值"""
        date_dt = pd.to_datetime(date).date()
        atr_row = self.daily_data[self.daily_data['trade_date'].dt.date == date_dt]
        if not atr_row.empty:
            return atr_row.iloc[0]['atr']
        return None
    
    def run_strategy(self):
        """运行ORB ATR策略"""
        print(f"开始运行 {self.symbol} ORB ATR策略 ({self.start_date} 至 {self.end_date})...")
        
        # 按交易日分组处理数据
        trade_dates = self.data['trade_date'].unique()
        
        for date in trade_dates:
            # 获取当天的ATR值
            atr = self._get_atr_for_date(date)
            if atr is None:
                print(f"警告: {date} ATR数据不可用，跳过")
                continue
                
            # 获取当天的数据
            day_data = self.data[self.data['trade_date'] == date].sort_values('date')
            
            # 确保当天至少有两根K线
            if len(day_data) < 2:
                print(f"警告: {date} 数据不足，跳过")
                continue
            
            # 提取前两根5分钟K线
            first_candle = day_data.iloc[0]
            second_candle = day_data.iloc[1]
            
            # 跳过十字星日
            if abs(first_candle['close'] - first_candle['open']) < 0.0001:
                print(f"{date}: 第一根K线为十字星，跳过交易")
                continue
            
            # 确定交易方向
            direction = 1 if first_candle['close'] > first_candle['open'] else -1
            
            # 入场价格 (第二根5分钟K线开盘价)
            entry_price = second_candle['open']
            
            # 使用ATR计算止损距离
            atr_stop_distance = self.atr_multiplier * atr
            
            # 设置止损价格 (ATR动态止损)
            stop_loss_price = entry_price - (atr_stop_distance * direction)
            
            # 计算风险
            risk = abs(entry_price - stop_loss_price)
            if risk < 0.0001:  # 防止除以零
                print(f"{date}: 风险太小，跳过交易")
                continue
            
            # 计算持仓规模
            shares = self._calculate_position_size(entry_price, stop_loss_price)
            
            # 如果持仓规模为0，跳过此次交易
            if shares == 0:
                print(f"{date}: 计算持仓为0，跳过交易")
                continue
            
            # 计算佣金
            commission = shares * self.commission_per_share
            
            # 提取当天剩余K线数据，用于模拟交易过程
            remaining_candles = day_data.iloc[2:]
            
            # 初始化出场变量
            exit_price = None
            exit_reason = None
            exit_time = None
            
            # 遍历剩余K线，检查是否触及止损
            for _, candle in remaining_candles.iterrows():
                # 做多情况
                if direction == 1:
                    # 检查是否触及止损
                    if candle['low'] <= stop_loss_price:
                        exit_price = stop_loss_price
                        exit_reason = 'Stop Loss'
                        exit_time = candle['date']
                        break
                # 做空情况
                else:
                    # 检查是否触及止损
                    if candle['high'] >= stop_loss_price:
                        exit_price = stop_loss_price
                        exit_reason = 'Stop Loss'
                        exit_time = candle['date']
                        break
            
            # 如果没有触及止损，则以收盘价平仓
            if exit_price is None:
                last_candle = day_data.iloc[-1]
                exit_price = last_candle['close']
                exit_reason = 'End of Day'
                exit_time = last_candle['date']
            
            # 计算交易盈亏
            trade_pnl = (exit_price - entry_price) * direction * shares - commission
            pnl_pct = trade_pnl / self.current_capital
            pnl_in_r = ((exit_price - entry_price) * direction) / risk
            
            # 记录交易
            trade = {
                'date': day_data.iloc[0]['date'].date(),
                'entry_time': second_candle['date'],
                'exit_time': exit_time,
                'direction': direction,
                'entry_price': entry_price,
                'stop_loss': stop_loss_price,
                'atr': atr,
                'atr_stop_distance': atr_stop_distance,
                'exit_price': exit_price,
                'exit_reason': exit_reason,
                'shares': shares,
                'risk_amount': self.current_capital * self.risk_per_trade,
                'risk_in_points': risk,
                'commission': commission,
                'pnl': trade_pnl,
                'pnl_pct': pnl_pct,
                'pnl_in_r': pnl_in_r
            }
            
            # 更新账户资金
            self.current_capital += trade_pnl
            
            # 添加交易记录
            self.trades.append(trade)
            
            # 打印交易摘要
            trade_direction = "多" if direction == 1 else "空"
            print(f"{date}: {trade_direction}头交易, 止损: ATR={atr:.4f}, 距离={atr_stop_distance:.4f}, 出场原因: {exit_reason}, 盈亏: {pnl_in_r:.2f}R (${trade_pnl:.2f})")
        
        # 计算交易结果
        self._calculate_trading_results()
        
        return self.trading_results
    
    def _calculate_position_size(self, entry_price, stop_loss_price):
        """
        计算持仓规模，根据公式:
        Shares = int[min((A * 0.01 / R), (4 * A / P))]
        
        A: 账户资金
        R: 入场价与止损价的差距
        P: 入场价格
        """
        risk = abs(entry_price - stop_loss_price)
        risk_amount = self.current_capital * self.risk_per_trade
        
        # 基于风险的头寸大小
        risk_based_shares = int(risk_amount / risk)
        
        # 基于杠杆的头寸大小
        leverage_based_shares = int((self.max_leverage * self.current_capital) / entry_price)
        
        # 取两者的较小值
        shares = min(risk_based_shares, leverage_based_shares)
        
        return shares
    
    def _calculate_trading_results(self):
        """计算交易结果统计数据"""
        if not self.trades:
            self.trading_results = {
                'symbol': self.symbol,
                'total_trades': 0,
                'profitable_trades': 0,
                'win_rate': 0,
                'total_pnl': 0,
                'total_pnl_pct': 0,
                'avg_pnl_in_r': 0,
                'initial_capital': self.initial_capital,
                'final_capital': self.current_capital,
                'trades': []
            }
            return
        
        # 转换交易记录为DataFrame以便分析
        trades_df = pd.DataFrame(self.trades)
        
        # 计算基本统计数据
        total_trades = len(trades_df)
        profitable_trades = len(trades_df[trades_df['pnl'] > 0])
        win_rate = profitable_trades / total_trades
        
        total_pnl = sum(trade['pnl'] for trade in self.trades)
        total_pnl_pct = (self.current_capital - self.initial_capital) / self.initial_capital
        
        # 按出场原因分组统计
        exit_reason_stats = trades_df.groupby('exit_reason').agg({
            'pnl': ['count', 'sum', 'mean'],
            'pnl_in_r': 'mean'
        })
        
        # 存储交易结果
        self.trading_results = {
            'symbol': self.symbol,
            'total_trades': total_trades,
            'profitable_trades': profitable_trades,
            'win_rate': win_rate,
            'total_pnl': total_pnl,
            'total_pnl_pct': total_pnl_pct,
            'avg_pnl_in_r': trades_df['pnl_in_r'].mean(),
            'exit_reason_stats': exit_reason_stats,
            'initial_capital': self.initial_capital,
            'final_capital': self.current_capital,
            'trades': trades_df
        }
        
        return self.trading_results
    
    def generate_excel_report(self, save_path=None):
        """将交易结果导出为Excel表格"""
        if not self.trading_results:
            print("没有交易结果，请先运行策略")
            return
        
        if save_path is None:
            output_dir = f"{config.OUTPUT_DIR}/{self.symbol}_ATR"
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)
            save_path = f"{output_dir}/{self.symbol}_ORB_ATR_Trading_Results.xlsx"
        
        print(f"正在生成Excel报告: {save_path}...")
        
        # 创建Excel工作簿和工作表
        wb = openpyxl.Workbook()
        
        # 创建摘要工作表
        summary_sheet = wb.active
        summary_sheet.title = "Trading Summary"
        
        # 添加标题
        summary_sheet['A1'] = f"{self.symbol} ORB ATR策略结果 (ATR: {self.atr_period}日, 乘数: {self.atr_multiplier})"
        summary_sheet['A1'].font = Font(bold=True, size=14)
        summary_sheet.merge_cells('A1:F1')
        
        # 添加基本信息
        summary_sheet['A3'] = "交易周期:"
        summary_sheet['B3'] = f"{self.start_date} 至 {self.end_date}"
        
        summary_sheet['A4'] = "初始资金:"
        summary_sheet['B4'] = f"${self.initial_capital:,.2f}"
        
        summary_sheet['A5'] = "最终资金:"
        summary_sheet['B5'] = f"${self.trading_results['final_capital']:,.2f}"
        
        summary_sheet['A6'] = "总盈亏:"
        summary_sheet['B6'] = f"${self.trading_results['total_pnl']:,.2f}"
        summary_sheet['C6'] = f"({self.trading_results['total_pnl_pct']:.2%})"
        
        # 添加交易统计
        summary_sheet['A8'] = "交易统计"
        summary_sheet['A8'].font = Font(bold=True, size=12)
        summary_sheet.merge_cells('A8:F8')
        
        summary_sheet['A9'] = "总交易次数:"
        summary_sheet['B9'] = self.trading_results['total_trades']
        
        summary_sheet['A10'] = "盈利交易:"
        summary_sheet['B10'] = self.trading_results['profitable_trades']
        
        summary_sheet['A11'] = "胜率:"
        summary_sheet['B11'] = f"{self.trading_results['win_rate']:.2%}"
        
        summary_sheet['A12'] = "平均R倍数:"
        summary_sheet['B12'] = f"{self.trading_results['avg_pnl_in_r']:.2f}R"
        
        # 按出场原因分析
        summary_sheet['A14'] = "出场原因分析"
        summary_sheet['A14'].font = Font(bold=True, size=12)
        summary_sheet.merge_cells('A14:F14')
        
        summary_sheet['A15'] = "出场原因"
        summary_sheet['B15'] = "次数"
        summary_sheet['C15'] = "总盈亏"
        summary_sheet['D15'] = "平均盈亏"
        summary_sheet['E15'] = "平均R倍数"
        
        # 设置标题行样式
        for col in range(1, 6):
            cell = summary_sheet.cell(row=15, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # 添加出场原因统计数据
        row = 16
        for exit_reason, stats in self.trading_results['exit_reason_stats'].iterrows():
            summary_sheet[f'A{row}'] = exit_reason
            summary_sheet[f'B{row}'] = int(stats[('pnl', 'count')])
            summary_sheet[f'C{row}'] = f"${stats[('pnl', 'sum')]:,.2f}"
            summary_sheet[f'D{row}'] = f"${stats[('pnl', 'mean')]:,.2f}"
            summary_sheet[f'E{row}'] = f"{stats[('pnl_in_r', 'mean')]:,.2f}R"
            row += 1
        
        # 创建详细交易记录工作表
        trades_sheet = wb.create_sheet("Trade Details")
        
        # 添加标题行
        headers = [
            "日期", "入场时间", "出场时间", "方向", 
            "入场价格", "止损价格", "ATR", "ATR止损距离", "出场价格", 
            "出场原因", "持仓数量", "风险金额($)", "风险距离(点)", 
            "佣金", "盈亏($)", "盈亏(%)", "盈亏(R)"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = trades_sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # 添加交易数据
        trades_data = self.trading_results['trades']
        
        for row_idx, trade in enumerate(trades_data.itertuples(), 2):
            # 日期和时间
            trades_sheet.cell(row=row_idx, column=1).value = trade.date.strftime('%Y-%m-%d')
            trades_sheet.cell(row=row_idx, column=2).value = pd.to_datetime(trade.entry_time).strftime('%H:%M:%S')
            trades_sheet.cell(row=row_idx, column=3).value = pd.to_datetime(trade.exit_time).strftime('%H:%M:%S')
            
            # 方向
            direction_text = "多" if trade.direction == 1 else "空"
            trades_sheet.cell(row=row_idx, column=4).value = direction_text
            
            # 价格
            trades_sheet.cell(row=row_idx, column=5).value = trade.entry_price
            trades_sheet.cell(row=row_idx, column=6).value = trade.stop_loss
            trades_sheet.cell(row=row_idx, column=7).value = trade.atr
            trades_sheet.cell(row=row_idx, column=8).value = trade.atr_stop_distance
            trades_sheet.cell(row=row_idx, column=9).value = trade.exit_price
            
            # 其他信息
            trades_sheet.cell(row=row_idx, column=10).value = trade.exit_reason
            trades_sheet.cell(row=row_idx, column=11).value = trade.shares
            trades_sheet.cell(row=row_idx, column=12).value = trade.risk_amount
            trades_sheet.cell(row=row_idx, column=13).value = trade.risk_in_points
            trades_sheet.cell(row=row_idx, column=14).value = trade.commission
            
            # 盈亏
            trades_sheet.cell(row=row_idx, column=15).value = trade.pnl
            trades_sheet.cell(row=row_idx, column=16).value = trade.pnl_pct
            trades_sheet.cell(row=row_idx, column=17).value = trade.pnl_in_r
            
            # 根据盈亏设置颜色
            if trade.pnl > 0:
                for col in range(1, 18):
                    trades_sheet.cell(row=row_idx, column=col).fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
            elif trade.pnl < 0:
                for col in range(1, 18):
                    trades_sheet.cell(row=row_idx, column=col).fill = PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid")
        
        # 创建权益曲线工作表
        equity_sheet = wb.create_sheet("Equity Curve")
        
        # 生成权益曲线数据
        equity_data = self._generate_equity_curve_data()
        
        # 添加标题行
        equity_sheet['A1'] = "日期"
        equity_sheet['B1'] = "账户价值"
        equity_sheet['C1'] = "累计回报率"
        
        # 设置标题行样式
        for col in ['A1', 'B1', 'C1']:
            equity_sheet[col].font = Font(bold=True)
            equity_sheet[col].fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # 添加权益曲线数据
        for i, (date, row) in enumerate(equity_data.iterrows(), 2):
            equity_sheet[f'A{i}'] = date.strftime('%Y-%m-%d')
            equity_sheet[f'B{i}'] = row['equity']
            equity_sheet[f'C{i}'] = (row['equity'] / self.initial_capital) - 1
            equity_sheet[f'C{i}'].number_format = '0.00%'
        
        # 设置列宽
        for sheet in [summary_sheet, trades_sheet, equity_sheet]:
            for col in range(1, sheet.max_column + 1):
                sheet.column_dimensions[get_column_letter(col)].width = 15
        
        # 保存Excel文件
        wb.save(save_path)
        print(f"Excel报告已保存到: {save_path}")
        
        return save_path
    
    def _generate_equity_curve_data(self):
        """生成权益曲线数据"""
        if not self.trading_results or len(self.trading_results['trades']) == 0:
            return pd.DataFrame()
        
        # 创建交易记录的副本
        trades_df = self.trading_results['trades'].copy()
        
        # 将日期转换为datetime
        trades_df['date'] = pd.to_datetime(trades_df['date'])
        trades_df = trades_df.sort_values('date')
        
        # 创建完整的日期范围
        first_date = trades_df['date'].min()
        last_date = trades_df['date'].max()
        
        # 创建日期索引
        date_range = pd.date_range(start=first_date, end=last_date)
        equity_df = pd.DataFrame(index=date_range)
        equity_df.index.name = 'date'
        
        # 初始化权益
        initial_equity = self.initial_capital
        current_equity = initial_equity
        
        # 填充每日权益
        for date, group in trades_df.groupby('date'):
            date_pnl = group['pnl'].sum()
            current_equity += date_pnl
            equity_df.loc[date, 'equity'] = current_equity
        
        # 填充缺失值
        equity_df['equity'] = equity_df['equity'].fillna(method='ffill')
        equity_df.loc[equity_df['equity'].isna(), 'equity'] = initial_equity
        
        return equity_df
    
    def generate_summary_report(self, save_path=None):
        """生成摘要统计报告"""
        if not self.trading_results:
            print("没有交易结果，请先运行策略")
            return None
        
        if save_path is None:
            output_dir = f"{config.OUTPUT_DIR}/{self.symbol}_ATR"
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)
            save_path = f"{output_dir}/{self.symbol}_ORB_ATR_Summary.txt"
        
        print(f"生成摘要报告: {save_path}")
        
        # 构建报告内容
        report = []
        report.append(f"===== {self.symbol} ORB ATR策略回测摘要 =====")
        report.append(f"策略: ORB ATR策略 (ATR周期: {self.atr_period}日, 乘数: {self.atr_multiplier})")
        report.append(f"交易周期: {self.start_date} 至 {self.end_date}")
        report.append(f"初始资金: ${self.initial_capital:,.2f}")
        report.append(f"最终资金: ${self.trading_results['final_capital']:,.2f}")
        report.append(f"总盈亏: ${self.trading_results['total_pnl']:,.2f} ({self.trading_results['total_pnl_pct']:.2%})")
        report.append("")
        
        report.append("--- 交易统计 ---")
        report.append(f"总交易次数: {self.trading_results['total_trades']}")
        report.append(f"盈利交易数: {self.trading_results['profitable_trades']}")
        report.append(f"亏损交易数: {self.trading_results['total_trades'] - self.trading_results['profitable_trades']}")
        report.append(f"胜率: {self.trading_results['win_rate']:.2%}")
        report.append(f"平均R倍数: {self.trading_results['avg_pnl_in_r']:.2f}R")
        report.append("")
        
        report.append("--- 出场原因分析 ---")
        for exit_reason, stats in self.trading_results['exit_reason_stats'].iterrows():
            count = int(stats[('pnl', 'count')])
            sum_pnl = stats[('pnl', 'sum')]
            avg_pnl = stats[('pnl', 'mean')]
            avg_r = stats[('pnl_in_r', 'mean')]
            report.append(f"{exit_reason}: {count}次, 总盈亏: ${sum_pnl:,.2f}, 平均盈亏: ${avg_pnl:,.2f}, 平均R倍数: {avg_r:.2f}R")
        
        # 将报告写入文件
        with open(save_path, 'w') as f:
            f.write('\n'.join(report))
        
        print(f"摘要报告已保存到: {save_path}")
        return '\n'.join(report)
    
    def plot_equity_curve(self, save_path=None, compare_with_original=False, original_results=None):
        """
        绘制权益曲线图
        
        参数:
        save_path: 图表保存路径，默认保存到输出目录
        compare_with_original: 是否与原始ORB策略比较
        original_results: 原始ORB策略的交易结果
        """
        if not self.trading_results:
            print("没有交易结果，请先运行策略")
            return
        
        # 生成权益曲线数据
        equity_df = self._generate_equity_curve_data()
        if equity_df.empty:
            print("没有足够的交易数据来绘制权益曲线")
            return
        
        plt.figure(figsize=(12, 8))
        
        # 绘制ATR策略权益曲线 - 修改为英文标签
        plt.plot(equity_df.index, equity_df['equity'], 'b-', linewidth=2, label='ORB ATR Strategy')
        
        # 计算真正的Buy & Hold策略
        try:
            # 获取标的资产的价格数据
            daily_data = self.daily_data.copy()
            daily_data = daily_data.sort_values('trade_date')
            
            # 计算起始价格和结束价格
            start_price = daily_data.iloc[0]['close']
            
            # 计算价格变化百分比
            daily_data['price_ratio'] = daily_data['close'] / start_price
            
            # 计算Buy & Hold策略的权益变化
            shares = self.initial_capital / start_price
            daily_data['buy_hold_equity'] = shares * daily_data['close']
            
            # 将Buy & Hold数据添加到equity_df
            buy_hold_df = daily_data[['trade_date', 'buy_hold_equity']].copy()
            buy_hold_df['date'] = buy_hold_df['trade_date']
            buy_hold_df = buy_hold_df.set_index('date')
            
            # 确保日期格式匹配
            buy_hold_df.index = pd.to_datetime(buy_hold_df.index)
            
            # 重采样到与equity_df相同的日期索引
            buy_hold_values = {}
            for date in equity_df.index:
                date_only = date.date()
                # 找到最接近的日期
                closest_dates = buy_hold_df.index[buy_hold_df.index.date <= date_only]
                if not closest_dates.empty:
                    closest_date = closest_dates[-1]
                    buy_hold_values[date] = buy_hold_df.loc[closest_date, 'buy_hold_equity']
                else:
                    buy_hold_values[date] = self.initial_capital
            
            # 创建Buy & Hold系列
            equity_df['buy_hold'] = pd.Series(buy_hold_values)
            
            # 绘制Buy & Hold曲线
            plt.plot(equity_df.index, equity_df['buy_hold'], 'g-', linewidth=1.5, label='Buy & Hold')
            
            print(f"Buy & Hold策略最终值: ${equity_df['buy_hold'].iloc[-1]:.2f}")
            
        except Exception as e:
            print(f"计算Buy & Hold策略时出错: {str(e)}")
            import traceback
            traceback.print_exc()
            
            # 如果计算失败，回退到简单水平线
            plt.axhline(y=self.initial_capital, color='g', linestyle='-', linewidth=1, label='Initial Capital')
        
        # 如果提供了原始ORB策略结果，也绘制其权益曲线用于比较
        if compare_with_original and original_results is not None:
            # 从原始结果生成权益曲线
            original_trades_df = original_results['trades'].copy()
            original_trades_df['date'] = pd.to_datetime(original_trades_df['date'])
            original_trades_df = original_trades_df.sort_values('date')
            
            # 创建完整的日期范围
            first_date = min(equity_df.index.min(), original_trades_df['date'].min())
            last_date = max(equity_df.index.max(), original_trades_df['date'].max())
            
            # 创建日期索引
            date_range = pd.date_range(start=first_date, end=last_date)
            original_equity_df = pd.DataFrame(index=date_range)
            original_equity_df.index.name = 'date'
            
            # 初始化权益
            original_equity = self.initial_capital
            
            # 填充每日权益
            for date, group in original_trades_df.groupby('date'):
                date_pnl = group['pnl'].sum()
                original_equity += date_pnl
                original_equity_df.loc[date, 'equity'] = original_equity
            
            # 填充缺失值
            original_equity_df['equity'] = original_equity_df['equity'].fillna(method='ffill')
            original_equity_df.loc[original_equity_df['equity'].isna(), 'equity'] = self.initial_capital
            
            # 绘制原始策略权益曲线 - 修改为英文标签
            plt.plot(original_equity_df.index, original_equity_df['equity'], 'r--', linewidth=2, label='Original ORB Strategy')
        
        # 设置图表标题和标签 - 修改为英文
        plt.title(f'{self.symbol} ORB ATR Strategy Equity Curve ({self.start_date} - {self.end_date})', fontsize=14)
        plt.xlabel('Date', fontsize=12)
        plt.ylabel('Portfolio Value ($)', fontsize=12)
        plt.grid(True, alpha=0.3)
        plt.legend()
        
        # 格式化X轴日期
        plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
        plt.gcf().autofmt_xdate()
        
        # 格式化Y轴为美元金额
        plt.gca().yaxis.set_major_formatter(mtick.StrMethodFormatter('${x:,.0f}'))
        
        # 添加最终权益标注 - 修改为英文
        final_equity = equity_df['equity'].iloc[-1]
        roi = (final_equity / self.initial_capital - 1) * 100
        plt.annotate(f'Final Equity: ${final_equity:,.2f} (ROI: {roi:.2f}%)',
                     xy=(equity_df.index[-1], final_equity),
                     xytext=(30, 20),
                     textcoords='offset points',
                     arrowprops=dict(arrowstyle='->', connectionstyle='arc3'),
                     bbox=dict(boxstyle='round,pad=0.5', fc='yellow', alpha=0.7))
        
        # 保存图表
        if save_path is None:
            output_dir = f"{config.OUTPUT_DIR}/{self.symbol}_ATR"
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)
            save_path = f"{output_dir}/{self.symbol}_ORB_ATR_Equity_Curve.png"
        
        plt.tight_layout()
        plt.savefig(save_path, dpi=300)
        print(f"权益曲线图已保存到: {save_path}")
        
        plt.close()


def run_atr_strategy(symbol, start_date=None, end_date=None, initial_capital=None, atr_period=14, atr_multiplier=0.05):
    """运行单个交易品种的ORB ATR策略"""
    try:
        print(f"\n===== 运行 {symbol} ORB ATR策略 =====")
        
        # 创建ATR策略实例
        strategy = ORB_ATR_Strategy(
            symbol=symbol,
            start_date=start_date,
            end_date=end_date,
            initial_capital=initial_capital,
            atr_period=atr_period,
            atr_multiplier=atr_multiplier
        )
        
        # 运行策略
        strategy.run_strategy()
        
        # 生成报告
        excel_path = strategy.generate_excel_report()
        summary = strategy.generate_summary_report()
        strategy.plot_equity_curve()
        
        print(f"\n===== {symbol} ORB ATR策略运行完成 =====")
        print(summary)
        
        return strategy.trading_results
        
    except Exception as e:
        print(f"运行 {symbol} ORB ATR策略时出错: {str(e)}")
        import traceback
        traceback.print_exc()
        return None


def run_all_symbols_atr():
    """运行所有配置的交易品种的ORB ATR策略"""
    results = {}
    for symbol in config.SYMBOLS:
        result = run_atr_strategy(symbol)
        results[symbol] = result
    
    return results


def compare_strategies(symbol, start_date=None, end_date=None, initial_capital=None, atr_period=14, atr_multiplier=0.05):
    """比较原始ORB策略和改进的ATR版本"""
    print(f"\n==== 比较 {symbol} 的原始ORB策略和ATR改进版 ====")
    
    # 导入原始ORB策略
    from ORB_Trading_Strategy import ORB_Trading_Strategy
    
    # 1. 运行原始ORB策略
    print(f"\n第1步: 运行原始ORB策略...")
    original_strategy = ORB_Trading_Strategy(
        symbol=symbol,
        start_date=start_date,
        end_date=end_date,
        initial_capital=initial_capital
    )
    original_results = original_strategy.run_strategy()
    
    # 2. 运行ATR改进版策略
    print(f"\n第2步: 运行ATR改进版ORB策略...")
    atr_strategy = ORB_ATR_Strategy(
        symbol=symbol,
        start_date=start_date,
        end_date=end_date,
        initial_capital=initial_capital,
        atr_period=atr_period,
        atr_multiplier=atr_multiplier
    )
    atr_results = atr_strategy.run_strategy()
    
    # 3. 比较结果
    print(f"\n第3步: 比较两种策略结果...")
    
    # 生成比较报告
    output_dir = f"{config.OUTPUT_DIR}/strategy_comparison"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    report_path = f"{output_dir}/{symbol}_strategy_comparison.txt"
    
    # 构建比较报告
    report = []
    report.append(f"===== {symbol} 策略对比报告 =====")
    report.append(f"交易周期: {start_date if start_date else config.START_DATE} 至 {end_date if end_date else config.END_DATE}")
    report.append(f"初始资金: ${initial_capital if initial_capital else config.INITIAL_CAPITAL:,.2f}")
    report.append("")
    
    report.append("--- 基本表现对比 ---")
    report.append("指标 | 原始ORB策略 | ATR改进版ORB策略")
    report.append("-----|------------|---------------")
    report.append(f"总交易次数 | {original_results['total_trades']} | {atr_results['total_trades']}")
    report.append(f"胜率 | {original_results['win_rate']:.2%} | {atr_results['win_rate']:.2%}")
    report.append(f"总盈亏 | ${original_results['total_pnl']:,.2f} | ${atr_results['total_pnl']:,.2f}")
    report.append(f"总收益率 | {original_results['total_pnl_pct']:.2%} | {atr_results['total_pnl_pct']:.2%}")
    report.append(f"平均R倍数 | {original_results['avg_pnl_in_r']:.2f}R | {atr_results['avg_pnl_in_r']:.2f}R")
    report.append("")
    
    report.append("--- 出场原因对比 ---")
    original_exits = original_results['exit_reason_stats']
    atr_exits = atr_results['exit_reason_stats']
    
    report.append("原始ORB策略:")
    for exit_reason, stats in original_exits.iterrows():
        count = int(stats[('pnl', 'count')])
        pct = count / original_results['total_trades'] * 100
        avg_r = stats[('pnl_in_r', 'mean')]
        report.append(f"- {exit_reason}: {count}次 ({pct:.1f}%), 平均R倍数: {avg_r:.2f}R")
    
    report.append("\nATR改进版ORB策略:")
    for exit_reason, stats in atr_exits.iterrows():
        count = int(stats[('pnl', 'count')])
        pct = count / atr_results['total_trades'] * 100
        avg_r = stats[('pnl_in_r', 'mean')]
        report.append(f"- {exit_reason}: {count}次 ({pct:.1f}%), 平均R倍数: {avg_r:.2f}R")
    
    # 保存比较报告
    with open(report_path, 'w') as f:
        f.write('\n'.join(report))
    
    print(f"策略对比报告已保存到: {report_path}")
    
    # 绘制权益曲线对比图
    equity_path = f"{output_dir}/{symbol}_equity_curve_comparison.png"
    atr_strategy.plot_equity_curve(equity_path, True, original_results)
    
    print("\n==== 策略比较完成 ====")
    print('\n'.join(report))
    
    return {
        'original': original_results,
        'atr': atr_results,
        'report_path': report_path,
        'equity_path': equity_path
    }


def main():
    """主函数"""
    import argparse
    
    parser = argparse.ArgumentParser(description='ORB ATR策略')
    parser.add_argument('--symbol', type=str, help='交易品种代码')
    parser.add_argument('--start_date', type=str, help='回测开始日期 (YYYY-MM-DD)')
    parser.add_argument('--end_date', type=str, help='回测结束日期 (YYYY-MM-DD)')
    parser.add_argument('--capital', type=float, help='初始资金')
    parser.add_argument('--atr_period', type=int, default=14, help='ATR计算周期')
    parser.add_argument('--atr_multiplier', type=float, default=0.05, help='ATR乘数')
    parser.add_argument('--compare', action='store_true', help='比较与原始ORB策略')
    parser.add_argument('--all', action='store_true', help='运行所有配置的交易品种')
    
    args = parser.parse_args()
    
    print(f"ORB ATR策略系统")
    print(f"交易周期: {args.start_date if args.start_date else config.START_DATE} 至 {args.end_date if args.end_date else config.END_DATE}")
    
    # 创建输出主目录
    if not os.path.exists(config.OUTPUT_DIR):
        os.makedirs(config.OUTPUT_DIR)
    
    if args.compare and args.symbol:
        # 比较单个品种的两种策略
        compare_strategies(
            symbol=args.symbol,
            start_date=args.start_date,
            end_date=args.end_date,
            initial_capital=args.capital,
            atr_period=args.atr_period,
            atr_multiplier=args.atr_multiplier
        )
    elif args.all:
        # 运行所有品种
        results = run_all_symbols_atr()
        print(f"\n所有ATR策略运行完成！详细结果已保存到 {config.OUTPUT_DIR} 目录")
    elif args.symbol:
        # 运行单个品种
        run_atr_strategy(
            symbol=args.symbol,
            start_date=args.start_date,
            end_date=args.end_date,
            initial_capital=args.capital,
            atr_period=args.atr_period,
            atr_multiplier=args.atr_multiplier
        )
    else:
        # 默认运行所有品种
        symbols = config.SYMBOLS
        for symbol in symbols:
            run_atr_strategy(
                symbol=symbol,
                start_date=args.start_date,
                end_date=args.end_date,
                initial_capital=args.capital,
                atr_period=args.atr_period,
                atr_multiplier=args.atr_multiplier
            )


if __name__ == "__main__":
    main()
